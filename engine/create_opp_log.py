from openpyxl import load_workbook
import sys
import time

def main():
    inputs = sys.argv[1] # input string
    inputDict = eval('dict('+inputs+')') # convert input string into dict
    create_opp_log(inputDict)
    print('Opportunity Log Created')

def create_opp_log(inputs):
        '''update quote log'''
        log = inputs['log']
        book = load_workbook(log)
        ws = book.worksheets[0]
        
        # get fist empty cell
        for cell in ws["A"]:
            if cell.value is None:
                current_row = cell.row
                break
        else:
            cell.row + 1

        # fix due date format
        due_date_list = inputs['due_date'].split('-') # break string into list
        due_date = f'{due_date_list[1]}/{due_date_list[2]}/{due_date_list[0][-2:]}' #relocate order

        # place inputs into spreadsheet
        ws[f'A{current_row}'] = inputs['quote_number']
        ws[f'B{current_row}'] = inputs['manager']
        ws[f'C{current_row}'] = time.strftime("%D")
        ws[f'D{current_row}'] = f'{inputs["project_name"]} {inputs["type_code"]}'
        ws[f'E{current_row}'] = inputs['project_location']
        ws[f'F{current_row}'] = due_date

        book.save(log)
        book.close()
    
if __name__ == "__main__":
    main()