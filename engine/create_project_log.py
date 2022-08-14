from openpyxl import load_workbook
import sys
import time

def main():
    inputs = ''
    inputs = sys.argv[1] # input string
    inputDict = eval('dict('+inputs+')') # convert input string into dict
    create_project_log(inputDict)
    print(f"{inputDict['p_sort']} Log Created")

def create_project_log(inputs):
    '''update project log'''
    log = inputs['log']
    book = load_workbook(log)

    if inputs['p_sort'] == 'project':
        ws = book.worksheets[0]
    elif inputs['p_sort'] == 'service':
        ws = book.worksheets[1]
    elif inputs['p_sort'] == 'HSE':
        ws = book.worksheets[2]
    
    # get fist empty cell
    for cell in ws["B"]:
        if cell.value is None:
            current_row = cell.row
            break
    else:
        cell.row + 1

    # place inputs into spreadsheet
    ws[f'A{current_row}'] = f"{inputs['p_project_number']}{inputs['p_project_billing']}"
    ws[f'B{current_row}'] = f"{inputs['p_project_state']}"
    ws[f'C{current_row}'] = f"{inputs['p_project_order']}"
    ws[f'D{current_row}'] = f"{inputs['p_project_name']} {inputs['p_type_code']}"
    ws[f'E{current_row}'] = f"{inputs['p_labor']}"
    ws[f'F{current_row}'] = time.strftime("%D")
    ws[f'G{current_row}'] = inputs['p_project_type']

    book.save(log)
    book.close()
    
if __name__ == "__main__":
    main()