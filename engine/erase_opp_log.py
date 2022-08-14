from openpyxl import load_workbook
import sys

def main():
    inputs = sys.argv[1] # input string
    inputDict = eval('dict('+inputs+')') # convert input string into dict
    erase_opp_log(inputDict)
    print('Opportunity Log Erased')

def erase_opp_log(inputs):
        '''erase last quote log entry'''
        log = inputs['log']
        book = load_workbook(log)
        ws = book.worksheets[0]
        
        # get fist empty cell
        for cell in ws["A"]:
            if cell.value is None:
                current_row = cell.row -1
                break
        else:
            cell.row + 1

        ws.delete_rows(current_row)

        book.save(log)
        book.close()
    
if __name__ == "__main__":
    main()